#!/usr/bin/env python3
"""
Personal Injury Attorney Directory - Data Import Script

Reads multiple Excel (.xlsx) and CSV (.csv) files with attorney data
and converts them into a structured JSON file used by the Astro static site generator.

Usage:
    python3 scripts/import.py

Input:  XLSX & CSV files in project root
Output: src/data/attorneys.json
"""

import csv
import json
import re
import sys
from datetime import date
from glob import glob
from pathlib import Path

import openpyxl

OUTPUT_FILE = Path("src") / "data" / "attorneys.json"

STATE_ABBREV_TO_SLUG = {
    "FL": "florida",
    "NJ": "new-jersey",
    "NY": "new-york",
    "CA": "california",
    "TX": "texas",
    "GA": "georgia",
    "PA": "pennsylvania",
    "IL": "illinois",
    "OH": "ohio",
    "NC": "north-carolina",
    "NM": "new-mexico",
    "UT": "utah",
    "CO": "colorado",
    "OK": "oklahoma",
    "AR": "arkansas",
    "RI": "rhode-island",
}

STATE_MAP = {
    "NJ": "NJ", "NEW JERSEY": "NJ",
    "FL": "FL", "FLORIDA": "FL",
    "NY": "NY", "NEW YORK": "NY",
    "TX": "TX", "TEXAS": "TX",
    "CA": "CA", "CALIFORNIA": "CA",
    "PA": "PA", "PENNSYLVANIA": "PA",
    "NM": "NM", "NEW MEXICO": "NM",
    "IL": "IL", "ILLINOIS": "IL",
    "UT": "UT", "UTAH": "UT",
    "NC": "NC", "NORTH CAROLINA": "NC",
    "RI": "RI", "RHODE ISLAND": "RI",
    "AR": "AR", "ARKANSAS": "AR",
    "CO": "CO", "COLORADO": "CO",
    "OK": "OK", "OKLAHOMA": "OK",
}


def slugify(text):
    """Convert any text to a URL-friendly lowercase slug."""
    if not text:
        return ""
    text = text.lower().strip()
    text = re.sub(r'\s*,\s*(esq|jr|sr|ii|iii)\.?\s*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[-_\s]+", "-", text)
    return text.strip("-")


def normalize_phone(phone):
    """Clean and standardize phone numbers."""
    if not phone:
        return ""
    phone = str(phone).strip()
    phone = re.sub(r"^1[-.\s]", "", phone)
    phone = re.sub(r"\s*-\s*", "-", phone)
    return phone


CITY_ALIASES = {
    "ft lauderdale": "Fort Lauderdale",
    "ft. lauderdale": "Fort Lauderdale",
    "ft myers": "Fort Myers",
    "ft. myers": "Fort Myers",
    "ft pierce": "Fort Pierce",
    "ft. pierce": "Fort Pierce",
    "ft walton beach": "Fort Walton Beach",
    "ft. walton beach": "Fort Walton Beach",
}


def normalize_city_name(city):
    """Collapse city name variants into a single canonical title-cased form."""
    if not city:
        return city

    city = re.sub(r"\s+", " ", city.strip())
    lookup_key = city.lower().rstrip(".")
    for key, canonical in CITY_ALIASES.items():
        if lookup_key == key.rstrip("."):
            return canonical

    if city.isupper() or city.islower():
        city = city.title()

    return city


def parse_address_robust(raw_address, default_state="FL"):
    """Extract city, state, and clean address string handling multi-location pipe-separated strings."""
    if not raw_address:
        return "", default_state, ""

    raw_address = re.sub(r'[\r\n\t]+', ' ', str(raw_address))
    parts = [p.strip() for p in raw_address.split('|') if p.strip()]
    target_part = parts[0]
    for part in parts:
        if re.search(r'\b(' + default_state + r'|New Jersey|Texas|Florida)\b', part, re.IGNORECASE):
            target_part = part
            break
            
    addr_clean = re.sub(r',\s*US\b', '', target_part, flags=re.IGNORECASE).strip()

    # Pattern 1: 'City, ST Personal Injury...'
    m = re.search(r'^\s*([A-Za-z\s\.\'-]+?),\s*(NJ|New Jersey|FL|Florida|NY|New York|TX|Texas|CA|California|PA|Pennsylvania)\s+Personal\s+Injury', addr_clean, re.IGNORECASE)
    if m:
        city = normalize_city_name(m.group(1).strip())
        st_raw = m.group(2).strip().upper()
        st = STATE_MAP.get(st_raw, default_state)
        return city, st, addr_clean

    # Pattern 1b: '..., ST, City, ZIP' e.g. '1599 Hamburg Turnpike, NJ, Wayne, 07470'
    m = re.search(r',\s*(NJ|New Jersey|FL|Florida|NY|New York|TX|Texas|CA|California|PA|Pennsylvania)\s*,\s*([A-Za-z\s\.\'-]+?)\s*,\s*(\d{5})', addr_clean, re.IGNORECASE)
    if m:
        st_raw = m.group(1).strip().upper()
        city = normalize_city_name(m.group(2).strip())
        st = STATE_MAP.get(st_raw, default_state)
        return city, st, addr_clean

    # Pattern 2: '..., City, ST ZIP' or '..., City, State ZIP'
    m = re.search(r',\s*([A-Za-z\s\.\'-]+?),\s*(NJ|New Jersey|FL|Florida|NY|New York|TX|Texas|CA|California|PA|Pennsylvania|NM|IL|UT|NC|RI|AR|CO|OK)\b(?:\s*,?\s*(\d{5}(-\d{4})?))?', addr_clean, re.IGNORECASE)
    if m:
        city = normalize_city_name(m.group(1).strip())
        st_raw = m.group(2).strip().upper()
        st = STATE_MAP.get(st_raw, default_state)
        return city, st, addr_clean

    # Pattern 3: '... Street City, ST ZIP' e.g. '48 South Street Morristown, New Jersey 07960'
    m = re.search(r'\s([A-Z][a-zA-Z\s]+(?:Township|Borough|City)?),\s*(NJ|FL|NY|CA|TX|New\s+Jersey|New\s+York|Florida)\s+\d{5}', addr_clean, re.IGNORECASE)
    if m:
        city = normalize_city_name(m.group(1).strip())
        st_raw = m.group(2).strip().upper()
        st = STATE_MAP.get(st_raw, default_state)
        return city, st, addr_clean

    # Pattern 4: 'City ST, ZIP' e.g. 'Freehold NJ, 07728'
    m = re.search(r'\b([A-Za-z\s\.\'-]+?)\s+(NJ|FL|NY|TX|CA|PA)\s*,\s*(\d{5})', addr_clean, re.IGNORECASE)
    if m:
        city = normalize_city_name(m.group(1).strip())
        st_raw = m.group(2).strip().upper()
        st = STATE_MAP.get(st_raw, default_state)
        return city, st, addr_clean

    # Pattern 5: 'City, ST' or '... City, ST' e.g. 'Newark, NJ' or 'Livingston, NJ'
    m = re.search(r'\b([A-Za-z\s\.\'-]+?),\s*(NJ|New Jersey|FL|Florida|NY|New York|TX|Texas|CA|California|PA|Pennsylvania)\s*$', addr_clean, re.IGNORECASE)
    if m:
        city_candidate = m.group(1).strip()
        words = city_candidate.split()
        city = ' '.join(words[-2:]) if len(words) > 2 else city_candidate
        city = normalize_city_name(city)
        st_raw = m.group(2).strip().upper()
        st = STATE_MAP.get(st_raw, default_state)
        return city, st, addr_clean

    # Pattern 6: 'Street City ST ZIP'
    m = re.search(r'\b([A-Z][a-zA-Z\s\.\'-]+?)\s+(NJ|FL|NY|TX|CA|PA|NM|IL|UT|NC|RI|AR|CO|OK)\s+(\d{5})', addr_clean, re.IGNORECASE)
    if m:
        city = normalize_city_name(m.group(1).strip())
        st_raw = m.group(2).strip().upper()
        st = STATE_MAP.get(st_raw, default_state)
        return city, st, addr_clean

    return "", default_state, addr_clean


NJ_AREA_CODES = {
    '201': 'Hackensack', '551': 'Jersey City',
    '973': 'Newark', '862': 'Newark',
    '732': 'Edison', '848': 'Woodbridge',
    '856': 'Cherry Hill',
    '609': 'Trenton', '640': 'Princeton',
}

TX_AREA_CODES = {
    '713': 'Houston', '281': 'Houston', '832': 'Houston', '346': 'Houston',
    '214': 'Dallas', '469': 'Dallas', '972': 'Dallas',
    '512': 'Austin', '737': 'Austin',
    '210': 'San Antonio', '726': 'San Antonio',
    '817': 'Fort Worth', '682': 'Fort Worth',
    '915': 'El Paso', '956': 'McAllen', '903': 'Tyler',
    '806': 'Lubbock', '361': 'Corpus Christi', '409': 'Beaumont'
}

FL_AREA_CODES = {
    '954': 'Fort Lauderdale', '754': 'Fort Lauderdale',
    '305': 'Miami', '786': 'Miami',
    '407': 'Orlando', '689': 'Orlando',
    '813': 'Tampa', '656': 'Tampa',
    '904': 'Jacksonville',
    '850': 'Tallahassee'
}


def resolve_attorney_location(row: dict, source_file: str, default_state: str):
    """Extract or infer city and state code for an attorney when explicit address is missing or incomplete."""
    address_raw = (row.get("Address") or row.get("address") or "").strip() if (row.get("Address") or row.get("address")) else ""
    
    city, state_code, clean_address = parse_address_robust(address_raw, default_state)
    if city:
        return city, state_code, clean_address

    # Check other fields for embedded address text (e.g. Office Phone column carrying address)
    for field in ["Office Phone", "Other Phones", "Phone", "phone", "firm", "Firm"]:
        val = str(row.get(field) or "").strip()
        c, s, ca = parse_address_robust(val, default_state)
        if c:
            return c, s, ca

    # Check filename e.g. 'Fort Lauderdale Area Data'
    if source_file:
        filename_match = re.match(r'([^/]+?)\s+Area\s+Data', source_file)
        if filename_match:
            return normalize_city_name(filename_match.group(1)), default_state, address_raw

    # Infer from phone area code
    phones = [row.get("Office Phone"), row.get("Other Phones"), row.get("Phone"), row.get("phone")]
    for p in phones:
        if not p:
            continue
        nums = re.sub(r"\D", "", str(p))
        if len(nums) >= 10:
            area = nums[-10:-7]
            if default_state == "NJ" and area in NJ_AREA_CODES:
                return NJ_AREA_CODES[area], "NJ", address_raw
            elif default_state == "TX" and area in TX_AREA_CODES:
                return TX_AREA_CODES[area], "TX", address_raw
            elif default_state == "FL" and area in FL_AREA_CODES:
                return FL_AREA_CODES[area], "FL", address_raw

    # Default fallback city per state
    fallback_city = "Newark" if default_state == "NJ" else ("Houston" if default_state == "TX" else "Miami")
    return fallback_city, default_state, address_raw


def parse_attorney_row(row: dict, index: int, source_file: str) -> dict | None:
    name = (row.get("Full Name") or row.get("Name") or row.get("name") or "").strip()
    nickname = (row.get("Nickname") or "").strip()
    bar_number = str(row.get("Bar Number") or "").strip()
    status = (row.get("Status") or "").strip()
    firm = (row.get("Firm") or row.get("firm") or "").strip()
    office_phone = normalize_phone(row.get("Office Phone") or row.get("Phone") or row.get("phone") or "")
    other_phones = (row.get("Other Phones") or "").strip()
    email = (row.get("Email") or row.get("email") or "").strip().lower()

    if not name:
        return None

    default_state = "NJ" if ("NJ" in source_file or "jersey" in source_file.lower()) else ("TX" if ("TX" in source_file or "texas" in source_file.lower()) else "FL")
    city, state_code, clean_address = resolve_attorney_location(row, source_file, default_state)

    state_slug = STATE_ABBREV_TO_SLUG.get(state_code, slugify(state_code) if state_code else "unknown")
    if state_slug == "unknown":
        state_slug = STATE_ABBREV_TO_SLUG.get(default_state, "florida")
        state_code = default_state

    city_slug = slugify(city)
    state_display = state_code if state_code else "USA"
    description = f"Personal injury attorney at {firm}" if firm else f"Personal injury attorney in {city}, {state_display}"

    attorney_slug = slugify(name)
    if not attorney_slug:
        attorney_slug = f"attorney-{index + 1}"

    return {
        "id": index + 1,
        "name": name,
        "nickname": nickname,
        "barNumber": bar_number,
        "status": status,
        "firm": firm,
        "state": state_slug,
        "stateCode": state_code,
        "city": city_slug,
        "cityDisplay": city,
        "address": clean_address,
        "phone": office_phone,
        "otherPhones": other_phones,
        "email": email,
        "description": description,
        "slug": attorney_slug,
    }


def process_xlsx_file(file_path: Path, start_index: int):
    print(f"\nProcessing {file_path.name}...")
    
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    headers = None
    for i in range(1, 6):
        row_values = [cell.value for cell in ws[i]]
        if 'Full Name' in row_values or 'Name' in row_values:
            headers = row_values
            header_row = i
            break
    
    if not headers:
        print(f"  Warning: No headers found in {file_path.name}, skipping")
        wb.close()
        return [], 0

    raw_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    total_raw = len(raw_rows)

    attorneys = []
    skipped = 0
    for i, row_data in enumerate(raw_rows):
        row_dict = dict(zip(headers, row_data))
        parsed = parse_attorney_row(row_dict, start_index + len(attorneys), file_path.name)
        if parsed:
            attorneys.append(parsed)
        else:
            skipped += 1

    wb.close()
    print(f"  Found {len(attorneys)} attorneys (skipped {skipped})")
    return attorneys, total_raw


def process_csv_file(file_path: Path, start_index: int):
    print(f"\nProcessing {file_path.name}...")
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = list(csv.DictReader(f))

    attorneys = []
    skipped = 0
    default_state = "NJ" if "new_jersey" in file_path.name.lower() else ("TX" if "texas" in file_path.name.lower() else "FL")

    for i, row in enumerate(reader):
        name = (row.get("Name") or row.get("name") or "").strip()
        if not name:
            skipped += 1
            continue

        phone = normalize_phone(row.get("Phone") or row.get("phone") or "")
        email = (row.get("Email") or row.get("email") or "").strip().lower()
        website = (row.get("Website") or row.get("website") or "").strip()
        linkedin = (row.get("LinkedIn") or row.get("linkedin") or "").strip()
        facebook = (row.get("Facebook") or row.get("facebook") or "").strip()
        firm = (row.get("Firm") or row.get("firm") or "").strip()

        city, state_code, clean_address = resolve_attorney_location(row, file_path.name, default_state)

        state_slug = STATE_ABBREV_TO_SLUG.get(state_code, slugify(state_code))
        city_slug = slugify(city)

        description = f"Personal injury attorney at {firm}" if firm else f"Personal injury attorney in {city}, {state_code}"
        attorney_slug = slugify(name)
        if not attorney_slug:
            attorney_slug = f"attorney-{start_index + len(attorneys) + 1}"

        attorneys.append({
            "id": start_index + len(attorneys) + 1,
            "name": name,
            "nickname": "",
            "barNumber": "",
            "status": "",
            "firm": firm,
            "state": state_slug,
            "stateCode": state_code,
            "city": city_slug,
            "cityDisplay": city,
            "address": clean_address,
            "phone": phone,
            "otherPhones": "",
            "email": email,
            "website": website,
            "linkedin": linkedin,
            "facebook": facebook,
            "description": description,
            "slug": attorney_slug,
        })

    print(f"  Found {len(attorneys)} attorneys (skipped {skipped})")
    return attorneys, len(reader)


def main():
    xlsx_files = sorted(glob("*.xlsx"))
    csv_files = sorted(glob("*.csv"))
    
    if not xlsx_files and not csv_files:
        print("Error: No XLSX or CSV files found")
        sys.exit(1)

    print(f"Found {len(xlsx_files)} XLSX files and {len(csv_files)} CSV files to process")

    all_attorneys = []
    total_raw_count = 0
    
    for file_path in xlsx_files:
        attorneys, raw_count = process_xlsx_file(Path(file_path), len(all_attorneys))
        all_attorneys.extend(attorneys)
        total_raw_count += raw_count

    for file_path in csv_files:
        attorneys, raw_count = process_csv_file(Path(file_path), len(all_attorneys))
        all_attorneys.extend(attorneys)
        total_raw_count += raw_count

    if not all_attorneys:
        print("\nError: No valid attorneys found after filtering.")
        sys.exit(1)

    # Deduplicate slugs
    seen_slugs = {}
    for attorney in all_attorneys:
        slug = attorney["slug"]
        if slug in seen_slugs:
            seen_slugs[slug] += 1
            attorney["slug"] = f"{slug}-{seen_slugs[slug]}"
        else:
            seen_slugs[slug] = 0

    cities = sorted(set(a["city"] for a in all_attorneys if a["city"]))
    states = sorted(set(a["state"] for a in all_attorneys if a["state"]))
    
    data = {
        "attorneys": all_attorneys,
        "metadata": {
            "total": len(all_attorneys),
            "totalRaw": total_raw_count,
            "skipped": total_raw_count - len(all_attorneys),
            "lastUpdated": date.today().isoformat(),
            "cities": cities,
            "states": states,
            "filesProcessed": len(xlsx_files) + len(csv_files),
        },
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\n✓ Imported {len(all_attorneys)} attorneys from {len(xlsx_files) + len(csv_files)} files")
    print(f"   Raw rows: {total_raw_count} | Skipped: {data['metadata']['skipped']} | Kept: {len(all_attorneys)}")
    print(f"   States: {', '.join(states)}")
    print(f"   Cities: {len(cities)}")
    print(f"   Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()

