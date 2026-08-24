import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# Create Workbook
wb = openpyxl.Workbook()

# Setup fonts and fills
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Arial", size=14, bold=True, color="1E293B")
subtitle_font = Font(name="Arial", size=10, italic=True, color="64748B")
bold_font = Font(name="Arial", size=10, bold=True)
regular_font = Font(name="Arial", size=10)

header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

high_severity_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Red
high_severity_font = Font(name="Arial", size=10, bold=True, color="991B1B")

med_severity_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Yellow/Orange
med_severity_font = Font(name="Arial", size=10, bold=True, color="92400E")

low_severity_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid") # Blue
low_severity_font = Font(name="Arial", size=10, bold=True, color="075985")

thin_border_side = Side(border_style="thin", color="CBD5E1")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

# ----------------------------------------------------
# 1. FINDINGS DATA
# ----------------------------------------------------
findings_data = [
    {
        "id": 1,
        "category": "1. ENTITY / CLINICAL BOUNDARY VIOLATIONS (CPOM)",
        "title": "CPOM — Unclarified Contingency Fee Disclosures & Expense Liability",
        "severity": "Medium",
        "ambiguous": "Yes",
        "file": "src/pages/index.astro",
        "lines": "65–69 (and FAQ Schema lines 65–69)",
        "excerpt": "Most personal injury attorneys work on a contingency fee basis - no upfront cost. They are paid only if you win, typically 33–40% of the settlement or court award.",
        "guideline": "Category 1 (Fee-splitting / outcome-contingent compensation disclosures) & Category 4 (UPL)",
        "why": "Mentions percentage-of-recovery fee structures (\"33–40% of the settlement or court award\") without the required state bar disclaimers clarifying that clients may remain responsible for court costs and case expenses regardless of legal outcome (e.g., Florida Bar Rule 4-1.5(f))."
    },
    {
        "id": 2,
        "category": "3. HIPAA / PRIVACY / PHI HANDLING",
        "title": "Privacy — Unencrypted Contact Intake Form Handling Potential Health/Injury Details Without Confidentiality / PHI Disclaimers",
        "severity": "High",
        "ambiguous": "No",
        "file": "src/pages/contact.astro",
        "lines": "52–54, 102–134",
        "excerpt": "Whether you're looking for an attorney, have a directory question, or want to list your practice - we'll get back to you within 1 business day.\n<form id=\"contact-form\" class=\"space-y-5\">",
        "guideline": "Category 3 (HIPAA / Privacy / PHI Handling)",
        "why": "Solicits injury-related user inquiries (\"Whether you're looking for an attorney...\") via a standard unencrypted web form without displaying a confidentiality warning, zero-PHI notice, or disclaimer that submitting information does not create an attorney-client relationship or guarantee HIPAA protection."
    },
    {
        "id": 3,
        "category": "4. UNAUTHORIZED PRACTICE OF LAW (UPL) / LEGAL CLAIMS",
        "title": "UPL — Direct Assessment of Entitlement Regarding Insurance Settlement Offers",
        "severity": "High",
        "ambiguous": "No",
        "file": "src/pages/when-to-hire-a-personal-injury-attorney.astro",
        "lines": "136–137",
        "excerpt": "Initial offers are almost always lower than what you are entitled to - often significantly so.",
        "guideline": "Category 4 (UPL / Legal Claims & Case Valuation) & Category 11 (General Red-Flag Language)",
        "why": "States a definitive legal conclusion regarding legal entitlement (\"lower than what you are entitled to\") for unrepresented users without an attorney evaluation step. Non-lawyer entities cannot evaluate legal entitlement or deliver case valuation opinions."
    },
    {
        "id": 4,
        "category": "4. UNAUTHORIZED PRACTICE OF LAW (UPL) / LEGAL CLAIMS",
        "title": "UPL — False Active Verification Status for Unverified Attorneys in Texas Template",
        "severity": "High",
        "ambiguous": "No",
        "file": "src/pages/personal-injury-attorneys/texas/[city]/[attorney].astro",
        "lines": "101–106",
        "excerpt": "<span class=\"inline-flex items-center gap-1 bg-green-50 text-green-700 text-xs font-semibold px-2.5 py-1 rounded-full border border-green-200\">\n  Licensed Attorney\n</span>",
        "guideline": "Category 4 (UPL / Credential Representation) & Category 5 (FTC Advertising Claims)",
        "why": "In the Texas attorney profile page, if attorney.status is missing/null, the component automatically falls back to rendering a green \"Licensed Attorney\" active status badge. Sourcing public records and falsely presenting unverified attorney entries as confirmed \"Licensed Attorney\" status misleads consumers and misrepresents bar standing."
    },
    {
        "id": 5,
        "category": "5. FTC / MARKETING / ADVERTISING CLAIMS",
        "title": "FTC — Absolute & Unsubstantiated Service Performance Claim (\"Handle Everything\")",
        "severity": "High",
        "ambiguous": "No",
        "file": "src/pages/index.astro",
        "lines": "144",
        "excerpt": "They handle everything - evidence gathering, insurance negotiations, and court litigation - so you can focus entirely on recovery while they fight for your compensation.",
        "guideline": "Category 5 (FTC / Marketing / Advertising Claims - Absolute Claims) & Category 11 (Red-Flag Language)",
        "why": "Uses absolute language (\"handle everything\") describing attorney services. Promising total coverage of all legal/factual needs is an unsubstantiated absolute performance claim under FTC rules (16 C.F.R. § 255) and state lawyer advertising rules."
    },
    {
        "id": 6,
        "category": "5. FTC / MARKETING / ADVERTISING CLAIMS",
        "title": "FTC — Unsubstantiated Comparative Performance Claim (\"Research Consistently Shows...\")",
        "severity": "High",
        "ambiguous": "No",
        "file": "src/pages/personal-injury-attorneys/index.astro",
        "lines": "107",
        "excerpt": "Research consistently shows that accident victims with legal representation receive significantly higher settlements than those who negotiate alone",
        "guideline": "Category 5 (FTC / Marketing Claims - Unsubstantiated Empirical Claims)",
        "why": "Asserts empirical research findings (\"Research consistently shows...\") regarding higher settlement outcomes without providing reference, study citation, or qualification required by FTC advertising guidelines."
    },
    {
        "id": 7,
        "category": "5. FTC / MARKETING / ADVERTISING CLAIMS",
        "title": "FTC — Contradictory & Inaccurate Geographical State Coverage Copy",
        "severity": "Medium",
        "ambiguous": "No",
        "file": "src/pages/index.astro",
        "lines": "87, 109, 115, 253",
        "excerpt": "Line 87: Browse licensed personal injury attorneys across the United States. Currently covering Florida - expanding nationwide.\nLine 109: Nationwide PI Attorney Directory - Active in Florida, New Jersey & Texas\nLine 115: Currently covering 54 cities across Florida, New Jersey, and Texas - expanding to all 50 states.\nLine 253: Currently covering Florida - with all 50 states coming.",
        "guideline": "Category 5 (FTC / Misleading Service Scope)",
        "why": "Contains conflicting statements on the homepage regarding active directory coverage (claiming FL-only in lines 87 & 253, but FL, NJ & TX in lines 109 & 115). Contradictory claims regarding service availability violate FTC truth-in-advertising principles."
    },
    {
        "id": 8,
        "category": "5. FTC / MARKETING / ADVERTISING CLAIMS",
        "title": "FTC — Outdated State Coverage Scope in Privacy Policy",
        "severity": "Medium",
        "ambiguous": "No",
        "file": "src/pages/privacy-policy.astro",
        "lines": "49, 96",
        "excerpt": "Line 49: FindPIAttorney.com ... is an independent online directory of personal injury attorneys licensed to practice in Florida and New Jersey.\nLine 96: ...maintained by state bar associations, including The Florida Bar and the New Jersey State Bar.",
        "guideline": "Category 5 (FTC / Accurate Disclosure Obligations)",
        "why": "The Privacy Policy explicitly limits its defined scope to Florida and New Jersey, omitting Texas. Texas attorney records and users are live on the site, making the legal privacy disclosures incomplete and inaccurate."
    },
    {
        "id": 9,
        "category": "5. FTC / MARKETING / ADVERTISING CLAIMS",
        "title": "FTC — Outdated State Coverage Scope in Terms of Service",
        "severity": "Medium",
        "ambiguous": "No",
        "file": "src/pages/terms-of-service.astro",
        "lines": "59",
        "excerpt": "Our current coverage includes Florida and New Jersey, with expansion to additional states ongoing.",
        "guideline": "Category 5 (FTC / Accurate Disclosure Obligations)",
        "why": "The Terms of Service state that active coverage is restricted to FL and NJ. Because TX is active in the codebase and live data, the scope representation in the governing contract is inaccurate."
    },
    {
        "id": 10,
        "category": "5. FTC / MARKETING / ADVERTISING CLAIMS",
        "title": "FTC — Outdated State Coverage Scope in Advertising Disclosure",
        "severity": "Medium",
        "ambiguous": "No",
        "file": "src/pages/advertising-disclosure.astro",
        "lines": "49, 61–62",
        "excerpt": "Line 49: ...listing personal injury attorneys licensed in Florida and New Jersey.\nLines 61–62: The Florida Bar ... The New Jersey State Bar Association...",
        "guideline": "Category 5 (FTC / Accurate Disclosure Obligations)",
        "why": "The Advertising Disclosure omits Texas State Bar regulatory references and Texas listing coverage despite live Texas directory pages and listings."
    },
    {
        "id": 11,
        "category": "5. FTC / MARKETING / ADVERTISING CLAIMS",
        "title": "FTC — Omission of Texas in Contact Page Information Card",
        "severity": "Low",
        "ambiguous": "No",
        "file": "src/pages/contact.astro",
        "lines": "91",
        "excerpt": "Coverage: Florida & New Jersey",
        "guideline": "Category 5 (FTC / Misleading Marketing Claims)",
        "why": "Displays a coverage badge stating \"Florida & New Jersey\", omitting Texas which is live in the directory."
    },
    {
        "id": 12,
        "category": "7. TCPA / CAN-SPAM / DIRECT COMMUNICATIONS",
        "title": "TCPA — Absence of Explicit Opt-in Consent Capture & Privacy Terms Link on Contact Form",
        "severity": "High",
        "ambiguous": "No",
        "file": "src/pages/contact.astro",
        "lines": "102–134",
        "excerpt": "<form id=\"contact-form\" class=\"space-y-5\">\n  ...\n  <button type=\"submit\" ...> Send Message </button>\n</form>",
        "guideline": "Category 7 (TCPA / Direct Communications Consent Capture)",
        "why": "The form collects user contact information (email, name) and triggers direct email communications without an explicit opt-in checkbox, legal consent language, or direct link to Privacy Policy/Terms near the submission button."
    },
    {
        "id": 13,
        "category": "8. AI GOVERNANCE / LABELING / TRANSPARENCY",
        "title": "AI Governance — Unsubstantiated \"Expert Advice\" Claim on Educational Content Placeholder",
        "severity": "Low",
        "ambiguous": "Yes",
        "file": "src/pages/index.astro",
        "lines": "265–266",
        "excerpt": "Legal Guides & Video Resources\nExpert advice, case studies, and what to expect when hiring a personal injury attorney.",
        "guideline": "Category 8 (AI Governance / Transparency) & Category 5 (FTC Marketing Claims)",
        "why": "Promotes upcoming content as \"Expert advice\" without identifying qualified human authors/attorneys or establishing an expert/AI review disclosure framework."
    },
    {
        "id": 14,
        "category": "9. EVIDENCE INTEGRITY / PROVENANCE",
        "title": "Evidence Integrity — Silent Fallback City Assignment & Data Transformation in Data Ingestion Pipeline",
        "severity": "Medium",
        "ambiguous": "Yes",
        "file": "scripts/import.py",
        "lines": "256, 368–373, 447–453",
        "excerpt": "fallback_city = \"Newark\" if default_state == \"NJ\" else (\"Houston\" if default_state == \"TX\" else \"Miami\")\ndescription = f\"Personal injury attorney at {firm}\" if firm else f\"Personal injury attorney in {city}, {state_display}\"",
        "guideline": "Category 9 (Evidence Integrity / Provenance)",
        "why": "When raw attorney address records lack a recognizable city, the import script automatically assigns default fallback cities (\"Newark\", \"Houston\", \"Miami\") and synthesizes attorney descriptions without flagging the data as inferred or transformed. Publishing incorrect location data for licensed professionals compromises data provenance."
    },
    {
        "id": 15,
        "category": "11. GENERAL PROHIBITED / RED-FLAG LANGUAGE",
        "title": "Red-Flag Language — Outcome Guarantees & Unqualified Contingency Claims",
        "severity": "Medium",
        "ambiguous": "Yes",
        "file": "src/pages/index.astro",
        "lines": "292",
        "excerpt": "They are paid only when you win, typically 33–40% of the settlement or verdict.",
        "guideline": "Category 11 (General Prohibited / Red-Flag Language - Certainty Language) & Category 4 (UPL)",
        "why": "Uses outcome-contingent phrasing (\"only when you win\") in a prominent FAQ block without an explicit disclaimer that past results do not guarantee future outcomes."
    }
]

# ----------------------------------------------------
# SHEET 1: Audit Findings (Main Table)
# ----------------------------------------------------
ws_findings = wb.active
ws_findings.title = "Audit Findings"
ws_findings.views.sheetView[0].showGridLines = True

# Title Header
ws_findings.cell(row=1, column=1, value="Compliance Audit Report — Lawyer Directory Findings").font = title_font
ws_findings.cell(row=2, column=1, value="Detailed list of all compliance, regulatory, UPL, FTC, TCPA, and privacy audit findings.").font = subtitle_font

headers_findings = [
    "Finding ID", "Category", "Finding Title", "Severity", "Ambiguous?", 
    "File Path", "Line Number(s)", "Excerpt / Code Snippet", "Guideline Violated", "Why It's a Violation"
]

for col_num, header_title in enumerate(headers_findings, 1):
    cell = ws_findings.cell(row=4, column=col_num, value=header_title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, item in enumerate(findings_data, 5):
    ws_findings.cell(row=row_idx, column=1, value=f"#{item['id']}").alignment = Alignment(horizontal="center", vertical="top")
    ws_findings.cell(row=row_idx, column=2, value=item['category']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_findings.cell(row=row_idx, column=3, value=item['title']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Severity styling
    sev_cell = ws_findings.cell(row=row_idx, column=4, value=item['severity'])
    sev_cell.alignment = Alignment(horizontal="center", vertical="top")
    if item['severity'] == "High":
        sev_cell.fill = high_severity_fill
        sev_cell.font = high_severity_font
    elif item['severity'] == "Medium":
        sev_cell.fill = med_severity_fill
        sev_cell.font = med_severity_font
    else:
        sev_cell.fill = low_severity_fill
        sev_cell.font = low_severity_font

    ws_findings.cell(row=row_idx, column=5, value=item['ambiguous']).alignment = Alignment(horizontal="center", vertical="top")
    ws_findings.cell(row=row_idx, column=6, value=item['file']).alignment = Alignment(horizontal="left", vertical="top")
    ws_findings.cell(row=row_idx, column=7, value=str(item['lines'])).alignment = Alignment(horizontal="center", vertical="top")
    ws_findings.cell(row=row_idx, column=8, value=item['excerpt']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_findings.cell(row=row_idx, column=9, value=item['guideline']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_findings.cell(row=row_idx, column=10, value=item['why']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Borders & fonts
    for c in range(1, 11):
        cell = ws_findings.cell(row=row_idx, column=c)
        cell.border = thin_border
        if c != 4: # don't overwrite severity font
            cell.font = regular_font

# Column Widths
col_widths_findings = {
    1: 12, 2: 30, 3: 35, 4: 14, 5: 14, 6: 35, 7: 18, 8: 45, 9: 30, 10: 50
}
for col_idx, width in col_widths_findings.items():
    ws_findings.column_dimensions[get_column_letter(col_idx)].width = width

ws_findings.freeze_panes = "A5"

# ----------------------------------------------------
# SHEET 2: Summary & Stats
# ----------------------------------------------------
ws_summary = wb.create_sheet(title="Executive Summary")
ws_summary.views.sheetView[0].showGridLines = True

ws_summary.cell(row=1, column=1, value="Compliance Audit — Executive Summary").font = title_font
ws_summary.cell(row=2, column=1, value="Report Generated: August 24, 2026 | Scope: Lawyer Directory Codebase").font = subtitle_font

# Overview Table
ws_summary.cell(row=4, column=1, value="Metric").font = header_font
ws_summary.cell(row=4, column=1).fill = header_fill
ws_summary.cell(row=4, column=2, value="Value").font = header_font
ws_summary.cell(row=4, column=2).fill = header_fill

summary_rows = [
    ("Total Audit Findings", len(findings_data)),
    ("Critical Severity Findings", 0),
    ("High Severity Findings", sum(1 for x in findings_data if x['severity'] == 'High')),
    ("Medium Severity Findings", sum(1 for x in findings_data if x['severity'] == 'Medium')),
    ("Low / Ambiguous Severity Findings", sum(1 for x in findings_data if x['severity'] == 'Low')),
]

for r_idx, (metric, val) in enumerate(summary_rows, 5):
    c1 = ws_summary.cell(row=r_idx, column=1, value=metric)
    c2 = ws_summary.cell(row=r_idx, column=2, value=val)
    c1.font = bold_font
    c2.font = bold_font
    c1.border = thin_border
    c2.border = thin_border
    c2.alignment = Alignment(horizontal="center")

# Category Breakdown Table
ws_summary.cell(row=12, column=1, value="Category").font = header_font
ws_summary.cell(row=12, column=1).fill = header_fill
ws_summary.cell(row=12, column=2, value="Finding Count").font = header_font
ws_summary.cell(row=12, column=2).fill = header_fill
ws_summary.cell(row=12, column=3, value="Status / Summary").font = header_font
ws_summary.cell(row=12, column=3).fill = header_fill

categories_list = [
    ("1. CPOM / Fee-Splitting & Contingency Disclosures", 1, "Unclarified expense liability in contingency disclosures"),
    ("2. FDA / SaMD Boundary", 0, "No issues found (directory strictly non-clinical)"),
    ("3. HIPAA / Privacy / PHI Handling", 1, "Unencrypted contact form lacking PHI/confidentiality notice"),
    ("4. UPL / Legal Claims & Case Valuation", 2, "Direct entitlement claim & false active bar status fallback"),
    ("5. FTC / Marketing & Advertising Claims", 7, "Absolute claims, empirical research claims & scope mismatches"),
    ("6. DSHEA / Supplement & Device Claims", 0, "No issues found"),
    ("7. TCPA / Direct Communications Consent", 1, "Absence of explicit opt-in consent capture on contact form"),
    ("8. AI Governance & Labeling", 1, "Unsubstantiated 'Expert Advice' claim on placeholder"),
    ("9. Evidence Integrity & Provenance", 1, "Silent fallback city assignment in import ETL script"),
    ("10. Insurance / Claims-Use Language", 0, "No issues found"),
    ("11. General Prohibited / Red-Flag Language", 1, "Outcome guarantees / certainty language without disclaimer")
]

for r_idx, (cat_name, cnt, note) in enumerate(categories_list, 13):
    c1 = ws_summary.cell(row=r_idx, column=1, value=cat_name)
    c2 = ws_summary.cell(row=r_idx, column=2, value=cnt)
    c3 = ws_summary.cell(row=r_idx, column=3, value=note)
    c1.font = regular_font
    c2.font = bold_font
    c3.font = regular_font
    c1.border = thin_border
    c2.border = thin_border
    c3.border = thin_border
    c2.alignment = Alignment(horizontal="center")

ws_summary.column_dimensions["A"].width = 45
ws_summary.column_dimensions["B"].width = 18
ws_summary.column_dimensions["C"].width = 65

# ----------------------------------------------------
# SHEET 3: Ambiguous Cases
# ----------------------------------------------------
ws_ambig = wb.create_sheet(title="Ambiguous & Borderline Cases")
ws_ambig.views.sheetView[0].showGridLines = True

ws_ambig.cell(row=1, column=1, value="Ambiguous / Borderline Compliance Cases").font = title_font
ws_ambig.cell(row=2, column=1, value="Findings that require legal counsel determination or policy decision.").font = subtitle_font

headers_ambig = ["Case #", "File & Lines", "Issue Description", "Legal Review & Uncertainty Details"]
for col_num, h_text in enumerate(headers_ambig, 1):
    c = ws_ambig.cell(row=4, column=col_num, value=h_text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

ambig_data = [
    (
        "Case 1",
        "src/pages/index.astro (Lines 65–69 & 292)",
        "Contingency Fee Descriptions & Outcome-Contingent Copy",
        "Explains that attorneys are paid '33–40% of the settlement or verdict' and 'only if you win'. As an informational directory (rather than a law firm), explaining general fee structures is informative. However, state bar rules (e.g. Florida Bar Rule 4-1.5(f)) require specific statements regarding client liability for costs when contingency fees are discussed publicly."
    ),
    (
        "Case 2",
        "scripts/import.py (Line 256)",
        "Synthetic Location Fallbacks in Data Pipeline",
        "Ingests Excel/CSV data and assigns fallback cities (e.g. 'Miami' or 'Houston') when addresses are unparseable. ETL script behavior is internal, but the resulting output populates public attorneys.json profiles with potentially inaccurate city associations. Legal review should determine if unverified location records should be omitted or flagged."
    ),
    (
        "Case 3",
        "README.md (Line 14) vs [attorney].astro",
        "Profile Contact Field Documentation Discrepancy",
        "README.md states profiles contain 'contact details', whereas state profile templates explicitly suppress phone, email, and street address. This is an internal developer documentation inconsistency."
    )
]

for r_idx, (c_id, file_loc, issue, details) in enumerate(ambig_data, 5):
    ws_ambig.cell(row=r_idx, column=1, value=c_id).alignment = Alignment(horizontal="center", vertical="top")
    ws_ambig.cell(row=r_idx, column=2, value=file_loc).alignment = Alignment(horizontal="left", vertical="top")
    ws_ambig.cell(row=r_idx, column=3, value=issue).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_ambig.cell(row=r_idx, column=4, value=details).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col in range(1, 5):
        cell = ws_ambig.cell(row=r_idx, column=col)
        cell.border = thin_border
        cell.font = regular_font

ws_ambig.column_dimensions["A"].width = 12
ws_ambig.column_dimensions["B"].width = 35
ws_ambig.column_dimensions["C"].width = 40
ws_ambig.column_dimensions["D"].width = 65
ws_ambig.freeze_panes = "A5"

# ----------------------------------------------------
# SHEET 4: Coverage & Scope
# ----------------------------------------------------
ws_scope = wb.create_sheet(title="Coverage & Scope Notes")
ws_scope.views.sheetView[0].showGridLines = True

ws_scope.cell(row=1, column=1, value="Audit Scope & Coverage Notes").font = title_font
ws_scope.cell(row=2, column=1, value="Scope limits and components audited.").font = subtitle_font

ws_scope.cell(row=4, column=1, value="Audit Area").font = header_font
ws_scope.cell(row=4, column=1).fill = header_fill
ws_scope.cell(row=4, column=2, value="Scope & Details").font = header_font
ws_scope.cell(row=4, column=2).fill = header_fill

scope_rows = [
    ("Fully Scanned Components", "All Astro pages (src/pages/*.astro and all state/city/attorney routes)"),
    ("Fully Scanned Layouts & Components", "All UI components (Header.astro, Footer.astro, AttorneyCard.astro, USAMapSVG.astro, etc.)"),
    ("Fully Scanned Legal Documents", "All legal documents (privacy-policy.astro, terms-of-service.astro, advertising-disclosure.astro)"),
    ("Fully Scanned Utilities & Schemas", "All helper utilities and data schemas (src/utils/attorneys.ts, src/types/attorney.ts)"),
    ("Fully Scanned Ingestion Scripts", "Data import and avatar generation scripts (scripts/import.py, scripts/generate-avatars.js)"),
    ("Fully Scanned Documentation", "Project documentation (README.md, package.json)"),
    ("Gaps / Structural Sampling", "src/data/attorneys.json contains 1,000+ auto-generated records. Individual attorney names/firms in attorneys.json were audited via structural sampling and import script logic rather than reading all 1,000+ JSON entries manually.")
]

for r_idx, (area, details) in enumerate(scope_rows, 5):
    c1 = ws_scope.cell(row=r_idx, column=1, value=area)
    c2 = ws_scope.cell(row=r_idx, column=2, value=details)
    c1.font = bold_font
    c2.font = regular_font
    c1.border = thin_border
    c2.border = thin_border
    c1.alignment = Alignment(vertical="top")
    c2.alignment = Alignment(vertical="top", wrap_text=True)

ws_scope.column_dimensions["A"].width = 35
ws_scope.column_dimensions["B"].width = 80

# Save Workbook
output_filename = "COMPLIANCE_AUDIT_REPORT.xlsx"
wb.save(output_filename)
print(f"Successfully generated {output_filename}")
